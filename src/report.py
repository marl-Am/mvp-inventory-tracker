from fpdf import FPDF
import os

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

os.makedirs("../reports", exist_ok=True)


def export_excel_report(data, output_dir, excel_filename="report.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Business Report"

    headers = [
        "Date",
        "Item",
        "Quantity",
        "Price",
        "Shipping Cost",
        "Expenses",
        "Income",
        "Revenue"
    ]
    ws.append(headers)

    for row in data:
        ws.append(row)

    column_widths = [20, 20, 12, 15, 15]
    for i, width in enumerate(column_widths, 1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = width

    excel_path = os.path.join(output_dir, excel_filename)
    wb.save(excel_path)
    print(f"Excel report created and saved to: {excel_path}")


def create_pdf_report(output_dir, report_dir, report_name="business_report.pdf"):
    os.makedirs(report_dir, exist_ok=True)

    charts = ["stock_by_item.png", "income_vs_expenses.png", "revenue_by_item.png"]

    pdf = FPDF()
    pdf.set_auto_page_break(True, margin=15)
    pdf.add_page()
    pdf.set_font("Arial", "B", size=18)
    pdf.cell(0, 10, f"Business Summary Report", ln=True, align="C")
    pdf.ln(10)

    for chart in charts:
        chart_path = os.path.join(output_dir, chart)
        if os.path.exists(chart_path):
            pdf.set_font("Arial", "B", size=12)
            pdf.cell(0, 10, chart.replace("_", " ").replace(".png", "").title(), ln=True)
            pdf.image(chart_path, w=180)
            pdf.ln(10)

        else:
            pdf.set_font("Arial", "I", size=10)
            pdf.cell(0, 10, f"Missing: {chart}", ln=True)
            pdf.ln(5)

    report_path = os.path.join(report_dir, report_name)
    pdf.output(report_path)
    print(f"PDF report created and saved to: {report_path}")

